from __future__ import annotations

import argparse
import csv
import re
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from auction_etl.database.session import engine


CLEAR_VALUE = "__CLEAR__"

CATALOG_PATTERNS = (
    re.compile(
        r"\b(?:MRZ?|DR|TACL|TATL|POCH|TRUE|UICZ|UPJY|"
        r"28TR|28MX|32TX|34TX|35TX|07TR|38TT|817|2427)"
        r"[-\s]?[0-9]{2,6}(?:[-/][0-9]{1,4})?\b",
        re.IGNORECASE,
    ),
    re.compile(
        r"\b[A-Z]{2,6}[-\s]?[0-9]{3,7}(?:[-/][0-9]{1,4})?\b",
        re.IGNORECASE,
    ),
)

REGION_PATTERNS = (
    (
        "Japan",
        re.compile(
            r"\bJapan(?:ese)?\b|日本盤|国内盤|日本国内|"
            r"昭和|Taurus|Tauras|トーラス",
            re.IGNORECASE,
        ),
    ),
    (
        "Taiwan",
        re.compile(
            r"\bTaiwan(?:ese)?\b|台湾盤|台灣盤|歌林|Kolin|Life",
            re.IGNORECASE,
        ),
    ),
    (
        "Hong Kong",
        re.compile(
            r"\bHong Kong\b|香港盤|香港版|港盤|港膠",
            re.IGNORECASE,
        ),
    ),
    (
        "Korea",
        re.compile(
            r"\bKorea(?:n)?\b|韓国盤|韓國盤",
            re.IGNORECASE,
        ),
    ),
    (
        "Malaysia",
        re.compile(
            r"\bMalaysia(?:n)?\b|馬來西亞|マレーシア",
            re.IGNORECASE,
        ),
    ),
    (
        "Singapore",
        re.compile(
            r"\bSingapore\b|新加坡|シンガポール",
            re.IGNORECASE,
        ),
    ),
    (
        "China",
        re.compile(
            r"\bChina\b|中國盤|中国盤|大陸盤",
            re.IGNORECASE,
        ),
    ),
)

MEDIA_RULES = (
    (
        "CD_BOX_SET",
        re.compile(
            r"\b(?:CD\s*)?(?:BOX|BOX SET)\b|CDボックス|"
            r"[0-9]+\s*(?:枚組|CD)\s*(?:CD)?BOX",
            re.IGNORECASE,
        ),
    ),
    (
        "CD_DVD_SET",
        re.compile(
            r"\bCD\s*/?\s*DVD\b|CD/DVD|DVD/CD",
            re.IGNORECASE,
        ),
    ),
    (
        "CASSETTE",
        re.compile(
            r"\bCASSETTE\b|カセット|テープ|TAPE",
            re.IGNORECASE,
        ),
    ),
    (
        "EP_7_INCH",
        re.compile(
            r"\bEP\b|7[\s-]*(?:INCH|インチ)|"
            r"7\"|シングルレコード|1x7",
            re.IGNORECASE,
        ),
    ),
    (
        "DVD",
        re.compile(
            r"\bDVD\b",
            re.IGNORECASE,
        ),
    ),
    (
        "CD",
        re.compile(
            r"\bCD\b|コンパクトディスク",
            re.IGNORECASE,
        ),
    ),
    (
        "LP",
        re.compile(
            r"\bLP\b|VINYL|レコード|33RPM",
            re.IGNORECASE,
        ),
    ),
)

DISC_COUNT_PATTERNS = (
    re.compile(
        r"\b([2-9]|[1-9][0-9]{1,2})\s*(?:LP|CD|DVD)\b",
        re.IGNORECASE,
    ),
    re.compile(
        r"\b([2-9]|[1-9][0-9]{1,2})\s*(?:枚|本)\s*(?:組|セット)",
        re.IGNORECASE,
    ),
    re.compile(
        r"\b([2-9]|[1-9][0-9]{1,2})[-\s]*(?:DISC|DISCS)\b",
        re.IGNORECASE,
    ),
)

BULK_PATTERNS = re.compile(
    r"まとめ|大量|セット|一括|bundle|lot|collection|"
    r"approximately\s+[0-9]+|[0-9]+\s*(?:枚|本|点|items?)",
    re.IGNORECASE,
)

OBI_PRESENT = re.compile(
    r"帯付|帯付き|帯あり|WITH OBI|OBI STRIP|OBI",
    re.IGNORECASE,
)
OBI_MISSING = re.compile(
    r"帯なし|帯無し|NO OBI|WITHOUT OBI",
    re.IGNORECASE,
)

INSERT_PRESENT = re.compile(
    r"INSERT(?: INCLUDED)?|ライナー|歌詞カード|"
    r"解説書|インサート|ブックレット",
    re.IGNORECASE,
)
INSERT_MISSING = re.compile(
    r"NO INSERT|WITHOUT INSERT|インサートなし|"
    r"歌詞カードなし|解説書なし",
    re.IGNORECASE,
)

POSTER_PRESENT = re.compile(
    r"POSTER|ポスター|PIN[-\s]?UP|ピンナップ",
    re.IGNORECASE,
)
POSTER_MISSING = re.compile(
    r"NO POSTER|WITHOUT POSTER|ポスターなし|"
    r"ピンナップなし",
    re.IGNORECASE,
)

RENTAL_PATTERN = re.compile(
    r"RENTAL|レンタル|貸出",
    re.IGNORECASE,
)
STICKER_PATTERN = re.compile(
    r"STICKER|ステッカー|シール",
    re.IGNORECASE,
)
PROMO_PATTERN = re.compile(
    r"PROMO|PROMOTIONAL|見本盤|サンプル盤|非売品",
    re.IGNORECASE,
)
SEALED_PATTERN = re.compile(
    r"SEALED|UNOPENED|未開封|新品",
    re.IGNORECASE,
)
REISSUE_PATTERN = re.compile(
    r"REISSUE|再発|復刻|180G|HEAVYWEIGHT|重量盤",
    re.IGNORECASE,
)
FIRST_PRESS_PATTERN = re.compile(
    r"FIRST PRESS|1ST PRESS|初回|初版|ORIGINAL PRESS",
    re.IGNORECASE,
)


@dataclass(slots=True)
class SellerStats:
    total_sales: int
    first_sale_at: datetime | None
    last_sale_at: datetime | None
    average_gross_price: Decimal | None
    repeat_seller: bool


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build and review collector-focused auction features."
    )
    subparsers = parser.add_subparsers(
        dest="command",
        required=True,
    )

    build_parser = subparsers.add_parser(
        "build",
        help="Generate automatic collector features.",
    )
    build_parser.add_argument(
        "--marketplace",
        choices=("buyee", "ebay", "all"),
        default="all",
    )
    build_parser.add_argument(
        "--force",
        action="store_true",
        help="Recalculate rows even when source data is unchanged.",
    )

    export_parser = subparsers.add_parser(
        "export",
        help="Export editable collector review tables.",
    )
    export_parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("review/collector"),
    )
    export_parser.add_argument(
        "--marketplace",
        choices=("buyee", "ebay", "all"),
        default="all",
    )

    import_parser = subparsers.add_parser(
        "import",
        help="Import collector review edits.",
    )
    import_parser.add_argument(
        "path",
        type=Path,
    )

    subparsers.add_parser(
        "status",
        help="Show collector-feature coverage.",
    )

    return parser.parse_args()


def money(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None

    try:
        return Decimal(str(value)).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )
    except (InvalidOperation, ValueError):
        return None


def normalize_catalog(value: str) -> str:
    return re.sub(
        r"\s+",
        "-",
        value.strip().upper(),
    )


def extract_catalog_number(text_value: str) -> str | None:
    candidates: list[str] = []

    for pattern in CATALOG_PATTERNS:
        candidates.extend(
            match.group(0)
            for match in pattern.finditer(text_value)
        )

    if not candidates:
        return None

    normalized = [
        normalize_catalog(candidate)
        for candidate in candidates
    ]

    normalized.sort(
        key=lambda value: (
            any(character.isalpha() for character in value),
            len(value),
        ),
        reverse=True,
    )

    return normalized[0]


def classify_region(text_value: str) -> str | None:
    for region, pattern in REGION_PATTERNS:
        if pattern.search(text_value):
            return region

    return None


def classify_media(
    text_value: str,
    existing_media: str | None,
) -> str | None:
    normalized_existing = (
        existing_media.strip().upper()
        if existing_media
        else None
    )

    for media_type, pattern in MEDIA_RULES:
        if pattern.search(text_value):
            return media_type

    return normalized_existing


def extract_disc_count(
    text_value: str,
    existing_count: int | None,
) -> int | None:
    counts: list[int] = []

    for pattern in DISC_COUNT_PATTERNS:
        counts.extend(
            int(match.group(1))
            for match in pattern.finditer(text_value)
        )

    if counts:
        return max(counts)

    return existing_count


def tri_state(
    text_value: str,
    present_pattern: re.Pattern[str],
    missing_pattern: re.Pattern[str],
) -> bool | None:
    if missing_pattern.search(text_value):
        return False

    if present_pattern.search(text_value):
        return True

    return None


def duration_days(
    started_at: datetime | None,
    ended_at: datetime | None,
) -> Decimal | None:
    if started_at is None or ended_at is None:
        return None

    seconds = Decimal(
        str((ended_at - started_at).total_seconds())
    )

    if seconds <= 0:
        return None

    return (
        seconds / Decimal("86400")
    ).quantize(
        Decimal("0.0001"),
        rounding=ROUND_HALF_UP,
    )


def safe_ratio(
    numerator: Decimal | None,
    denominator: Decimal | None,
) -> Decimal | None:
    if numerator is None or denominator is None:
        return None

    if denominator <= 0:
        return None

    return (
        numerator / denominator
    ).quantize(
        Decimal("0.0001"),
        rounding=ROUND_HALF_UP,
    )


def calculate_score(
    *,
    catalog_number: str | None,
    region: str | None,
    media_type: str | None,
    obi: bool | None,
    insert: bool | None,
    poster: bool | None,
    promo: bool,
    sealed: bool,
    first_press: bool,
    reissue: bool,
    rental: bool,
    bulk_lot: bool,
    bid_count: int | None,
    price_multiplier: Decimal | None,
    gross_price: Decimal | None,
) -> int:
    score = 20

    if catalog_number:
        score += 10

    if region in {
        "Hong Kong",
        "Taiwan",
        "Korea",
        "Malaysia",
        "Singapore",
    }:
        score += 8
    elif region == "Japan":
        score += 4

    if media_type in {
        "LP",
        "EP_7_INCH",
        "CASSETTE",
    }:
        score += 5

    if obi is True:
        score += 8
    elif obi is False:
        score -= 4

    if insert is True:
        score += 6
    elif insert is False:
        score -= 3

    if poster is True:
        score += 7

    if promo:
        score += 12

    if sealed:
        score += 8

    if first_press:
        score += 8

    if reissue:
        score -= 4

    if rental:
        score -= 8

    if bulk_lot:
        score -= 15

    if bid_count is not None:
        if bid_count >= 40:
            score += 10
        elif bid_count >= 20:
            score += 7
        elif bid_count >= 10:
            score += 4

    if price_multiplier is not None:
        if price_multiplier >= Decimal("20"):
            score += 10
        elif price_multiplier >= Decimal("10"):
            score += 7
        elif price_multiplier >= Decimal("5"):
            score += 4

    if gross_price is not None:
        if gross_price >= Decimal("30000"):
            score += 8
        elif gross_price >= Decimal("10000"):
            score += 5
        elif gross_price >= Decimal("5000"):
            score += 2

    return max(0, min(100, score))


def verdict_for_score(
    score: int,
    *,
    bulk_lot: bool,
    reissue: bool,
) -> str:
    if bulk_lot:
        return "BULK_REVIEW"

    if score >= 80:
        return "PRIORITY_TARGET"

    if score >= 65:
        return "STRONG_INTEREST"

    if score >= 50:
        return "WATCH"

    if reissue:
        return "LOW_PRIORITY_REISSUE"

    if score >= 35:
        return "REFERENCE_ONLY"

    return "PASS"


def create_schema() -> None:
    statements = (
        """
        CREATE TABLE IF NOT EXISTS warehouse.auction_collector (
            marketplace VARCHAR(32) NOT NULL,
            listing_id VARCHAR(128) NOT NULL,

            source_fingerprint TEXT,

            auto_catalog_number VARCHAR(128),
            manual_catalog_number VARCHAR(128),

            auto_region VARCHAR(64),
            manual_region VARCHAR(64),

            auto_media_type VARCHAR(64),
            manual_media_type VARCHAR(64),

            auto_disc_count INTEGER,
            manual_disc_count INTEGER,

            auto_bulk_lot BOOLEAN,
            manual_bulk_lot BOOLEAN,

            auto_obi BOOLEAN,
            manual_obi BOOLEAN,

            auto_insert_present BOOLEAN,
            manual_insert_present BOOLEAN,

            auto_poster_present BOOLEAN,
            manual_poster_present BOOLEAN,

            auto_rental BOOLEAN,
            manual_rental BOOLEAN,

            auto_sticker BOOLEAN,
            manual_sticker BOOLEAN,

            auto_promo BOOLEAN,
            manual_promo BOOLEAN,

            auto_sealed BOOLEAN,
            manual_sealed BOOLEAN,

            auto_reissue BOOLEAN,
            manual_reissue BOOLEAN,

            auto_first_press BOOLEAN,
            manual_first_press BOOLEAN,

            seller_total_sales INTEGER,
            seller_first_sale_at TIMESTAMPTZ,
            seller_last_sale_at TIMESTAMPTZ,
            seller_average_gross_price NUMERIC(18, 2),
            repeat_seller BOOLEAN,

            auction_duration_days NUMERIC(12, 4),
            start_to_finish_multiplier NUMERIC(18, 4),
            bids_per_day NUMERIC(18, 4),

            auto_importance_score INTEGER,
            manual_importance_score INTEGER,

            auto_verdict VARCHAR(64),
            manual_verdict VARCHAR(64),

            manual_condition_media VARCHAR(64),
            manual_condition_cover VARCHAR(64),
            manual_completeness_notes TEXT,
            manual_collector_notes TEXT,

            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),

            PRIMARY KEY (marketplace, listing_id)
        )
        """,
        """
        CREATE OR REPLACE VIEW warehouse.auction_collector_effective AS
        SELECT
            a.*,

            c.manual_catalog_number,
            c.manual_region,
            c.manual_media_type,
            c.manual_disc_count,
            c.manual_bulk_lot,
            c.manual_obi,
            c.manual_insert_present,
            c.manual_poster_present,
            c.manual_rental,
            c.manual_sticker,
            c.manual_promo,
            c.manual_sealed,
            c.manual_reissue,
            c.manual_first_press,
            c.manual_importance_score,
            c.manual_verdict,
            c.manual_condition_media,
            c.manual_condition_cover,
            c.manual_completeness_notes,
            c.manual_collector_notes,

            COALESCE(
                c.manual_catalog_number,
                c.auto_catalog_number,
                a.catalog_number
            ) AS effective_catalog_number,

            COALESCE(
                c.manual_region,
                c.auto_region
            ) AS effective_region,

            COALESCE(
                c.manual_media_type,
                c.auto_media_type,
                a.media_type
            ) AS effective_media_type,

            COALESCE(
                c.manual_disc_count,
                c.auto_disc_count,
                a.disc_count
            ) AS effective_disc_count,

            COALESCE(
                c.manual_bulk_lot,
                c.auto_bulk_lot,
                a.bulk_lot
            ) AS effective_bulk_lot,

            COALESCE(
                c.manual_obi,
                c.auto_obi
            ) AS effective_obi,

            COALESCE(
                c.manual_insert_present,
                c.auto_insert_present
            ) AS effective_insert_present,

            COALESCE(
                c.manual_poster_present,
                c.auto_poster_present
            ) AS effective_poster_present,

            COALESCE(
                c.manual_rental,
                c.auto_rental
            ) AS effective_rental,

            COALESCE(
                c.manual_sticker,
                c.auto_sticker
            ) AS effective_sticker,

            COALESCE(
                c.manual_promo,
                c.auto_promo
            ) AS effective_promo,

            COALESCE(
                c.manual_sealed,
                c.auto_sealed
            ) AS effective_sealed,

            COALESCE(
                c.manual_reissue,
                c.auto_reissue
            ) AS effective_reissue,

            COALESCE(
                c.manual_first_press,
                c.auto_first_press
            ) AS effective_first_press,

            c.seller_total_sales,
            c.seller_first_sale_at,
            c.seller_last_sale_at,
            c.seller_average_gross_price,
            c.repeat_seller,

            c.auction_duration_days,
            c.start_to_finish_multiplier,
            c.bids_per_day,

            COALESCE(
                c.manual_importance_score,
                c.auto_importance_score
            ) AS effective_importance_score,

            COALESCE(
                c.manual_verdict,
                c.auto_verdict
            ) AS effective_verdict,

            COALESCE(
                c.manual_condition_media,
                a.condition_media
            ) AS effective_condition_media,

            COALESCE(
                c.manual_condition_cover,
                a.condition_cover
            ) AS effective_condition_cover,

            c.updated_at AS collector_updated_at

        FROM warehouse.auction AS a
        LEFT JOIN warehouse.auction_collector AS c
          ON c.marketplace = a.marketplace
         AND c.listing_id = a.listing_id
        """,
    )

    with engine.begin() as connection:
        for statement in statements:
            connection.execute(text(statement))


def load_seller_stats() -> dict[tuple[str, str], SellerStats]:
    statement = text(
        """
        SELECT
            marketplace,
            COALESCE(NULLIF(BTRIM(seller), ''), '[UNKNOWN]')
                AS seller_key,
            COUNT(*) AS total_sales,
            MIN(ended_at) AS first_sale_at,
            MAX(ended_at) AS last_sale_at,
            AVG(gross_price) AS average_gross_price
        FROM warehouse.auction
        GROUP BY
            marketplace,
            COALESCE(NULLIF(BTRIM(seller), ''), '[UNKNOWN]')
        """
    )

    with engine.connect() as connection:
        rows = connection.execute(
            statement
        ).mappings().all()

    result: dict[tuple[str, str], SellerStats] = {}

    for row in rows:
        total_sales = int(row["total_sales"])

        result[
            (
                str(row["marketplace"]),
                str(row["seller_key"]),
            )
        ] = SellerStats(
            total_sales=total_sales,
            first_sale_at=row["first_sale_at"],
            last_sale_at=row["last_sale_at"],
            average_gross_price=money(
                row["average_gross_price"]
            ),
            repeat_seller=total_sales >= 2,
        )

    return result


def load_auctions(
    marketplace: str,
) -> list[dict[str, Any]]:
    where_clause = ""

    if marketplace != "all":
        where_clause = "WHERE a.marketplace = :marketplace"

    statement = text(
        f"""
        SELECT
            a.marketplace,
            a.listing_id,
            a.title,
            a.artist,
            a.seller,
            a.media_type,
            a.disc_count,
            a.bulk_lot,
            a.catalog_number,
            a.condition_media,
            a.condition_cover,
            a.bid_count,
            a.start_price,
            a.final_price,
            a.tax_amount,
            a.gross_price,
            a.currency,
            a.ended_at,
            d.started_at,
            d.condition_text,
            d.detail_price
        FROM warehouse.auction AS a
        LEFT JOIN warehouse.auction_detail AS d
          ON d.marketplace = a.marketplace
         AND d.listing_id = a.listing_id
        {where_clause}
        ORDER BY
            a.ended_at DESC NULLS LAST,
            a.id DESC
        """
    )

    parameters = (
        {}
        if marketplace == "all"
        else {"marketplace": marketplace}
    )

    with engine.connect() as connection:
        rows = connection.execute(
            statement,
            parameters,
        ).mappings().all()

    return [dict(row) for row in rows]


def fingerprint(row: dict[str, Any]) -> str:
    values = (
        row.get("title"),
        row.get("seller"),
        row.get("media_type"),
        row.get("disc_count"),
        row.get("bulk_lot"),
        row.get("catalog_number"),
        row.get("bid_count"),
        row.get("start_price"),
        row.get("gross_price"),
        row.get("started_at"),
        row.get("ended_at"),
        row.get("condition_text"),
    )

    return "|".join(
        "" if value is None else str(value)
        for value in values
    )


def build_features(
    marketplace: str,
    force: bool,
) -> tuple[int, int]:
    create_schema()

    seller_stats = load_seller_stats()
    rows = load_auctions(marketplace)

    existing_statement = text(
        """
        SELECT source_fingerprint
        FROM warehouse.auction_collector
        WHERE marketplace = :marketplace
          AND listing_id = :listing_id
        """
    )

    upsert_statement = text(
        """
        INSERT INTO warehouse.auction_collector (
            marketplace,
            listing_id,
            source_fingerprint,

            auto_catalog_number,
            auto_region,
            auto_media_type,
            auto_disc_count,
            auto_bulk_lot,

            auto_obi,
            auto_insert_present,
            auto_poster_present,
            auto_rental,
            auto_sticker,
            auto_promo,
            auto_sealed,
            auto_reissue,
            auto_first_press,

            seller_total_sales,
            seller_first_sale_at,
            seller_last_sale_at,
            seller_average_gross_price,
            repeat_seller,

            auction_duration_days,
            start_to_finish_multiplier,
            bids_per_day,

            auto_importance_score,
            auto_verdict,
            updated_at
        )
        VALUES (
            :marketplace,
            :listing_id,
            :source_fingerprint,

            :auto_catalog_number,
            :auto_region,
            :auto_media_type,
            :auto_disc_count,
            :auto_bulk_lot,

            :auto_obi,
            :auto_insert_present,
            :auto_poster_present,
            :auto_rental,
            :auto_sticker,
            :auto_promo,
            :auto_sealed,
            :auto_reissue,
            :auto_first_press,

            :seller_total_sales,
            :seller_first_sale_at,
            :seller_last_sale_at,
            :seller_average_gross_price,
            :repeat_seller,

            :auction_duration_days,
            :start_to_finish_multiplier,
            :bids_per_day,

            :auto_importance_score,
            :auto_verdict,
            NOW()
        )
        ON CONFLICT (marketplace, listing_id)
        DO UPDATE SET
            source_fingerprint =
                EXCLUDED.source_fingerprint,

            auto_catalog_number =
                EXCLUDED.auto_catalog_number,
            auto_region =
                EXCLUDED.auto_region,
            auto_media_type =
                EXCLUDED.auto_media_type,
            auto_disc_count =
                EXCLUDED.auto_disc_count,
            auto_bulk_lot =
                EXCLUDED.auto_bulk_lot,

            auto_obi =
                EXCLUDED.auto_obi,
            auto_insert_present =
                EXCLUDED.auto_insert_present,
            auto_poster_present =
                EXCLUDED.auto_poster_present,
            auto_rental =
                EXCLUDED.auto_rental,
            auto_sticker =
                EXCLUDED.auto_sticker,
            auto_promo =
                EXCLUDED.auto_promo,
            auto_sealed =
                EXCLUDED.auto_sealed,
            auto_reissue =
                EXCLUDED.auto_reissue,
            auto_first_press =
                EXCLUDED.auto_first_press,

            seller_total_sales =
                EXCLUDED.seller_total_sales,
            seller_first_sale_at =
                EXCLUDED.seller_first_sale_at,
            seller_last_sale_at =
                EXCLUDED.seller_last_sale_at,
            seller_average_gross_price =
                EXCLUDED.seller_average_gross_price,
            repeat_seller =
                EXCLUDED.repeat_seller,

            auction_duration_days =
                EXCLUDED.auction_duration_days,
            start_to_finish_multiplier =
                EXCLUDED.start_to_finish_multiplier,
            bids_per_day =
                EXCLUDED.bids_per_day,

            auto_importance_score =
                EXCLUDED.auto_importance_score,
            auto_verdict =
                EXCLUDED.auto_verdict,
            updated_at = NOW()
        """
    )

    scanned = 0
    changed = 0

    with engine.begin() as connection:
        for row in rows:
            scanned += 1

            row_fingerprint = fingerprint(row)

            existing = connection.execute(
                existing_statement,
                {
                    "marketplace": row["marketplace"],
                    "listing_id": row["listing_id"],
                },
            ).scalar_one_or_none()

            if (
                not force
                and existing == row_fingerprint
            ):
                continue

            combined_text = " ".join(
                str(value)
                for value in (
                    row.get("title"),
                    row.get("artist"),
                    row.get("condition_text"),
                )
                if value
            )

            catalog_number = (
                extract_catalog_number(combined_text)
                or row.get("catalog_number")
            )
            region = classify_region(combined_text)
            media_type = classify_media(
                combined_text,
                row.get("media_type"),
            )
            disc_count = extract_disc_count(
                combined_text,
                row.get("disc_count"),
            )

            bulk_lot = bool(
                row.get("bulk_lot")
                or BULK_PATTERNS.search(combined_text)
                or (
                    disc_count is not None
                    and disc_count >= 4
                )
            )

            obi = tri_state(
                combined_text,
                OBI_PRESENT,
                OBI_MISSING,
            )
            insert_present = tri_state(
                combined_text,
                INSERT_PRESENT,
                INSERT_MISSING,
            )
            poster_present = tri_state(
                combined_text,
                POSTER_PRESENT,
                POSTER_MISSING,
            )

            rental = bool(
                RENTAL_PATTERN.search(combined_text)
            )
            sticker = bool(
                STICKER_PATTERN.search(combined_text)
            )
            promo = bool(
                PROMO_PATTERN.search(combined_text)
            )
            sealed = bool(
                SEALED_PATTERN.search(combined_text)
            )
            reissue = bool(
                REISSUE_PATTERN.search(combined_text)
            )
            first_press = bool(
                FIRST_PRESS_PATTERN.search(combined_text)
            )

            started_at = row.get("started_at")
            ended_at = row.get("ended_at")
            duration = duration_days(
                started_at,
                ended_at,
            )

            start_price = money(
                row.get("start_price")
            )
            gross_price = money(
                row.get("gross_price")
            )

            price_multiplier = safe_ratio(
                gross_price,
                start_price,
            )

            bid_count = row.get("bid_count")
            bids_per_day = safe_ratio(
                (
                    Decimal(bid_count)
                    if bid_count is not None
                    else None
                ),
                duration,
            )

            seller_key = (
                str(row.get("seller")).strip()
                if row.get("seller")
                else "[UNKNOWN]"
            )

            seller = seller_stats.get(
                (
                    str(row["marketplace"]),
                    seller_key,
                ),
                SellerStats(
                    total_sales=1,
                    first_sale_at=ended_at,
                    last_sale_at=ended_at,
                    average_gross_price=gross_price,
                    repeat_seller=False,
                ),
            )

            score = calculate_score(
                catalog_number=catalog_number,
                region=region,
                media_type=media_type,
                obi=obi,
                insert=insert_present,
                poster=poster_present,
                promo=promo,
                sealed=sealed,
                first_press=first_press,
                reissue=reissue,
                rental=rental,
                bulk_lot=bulk_lot,
                bid_count=bid_count,
                price_multiplier=price_multiplier,
                gross_price=gross_price,
            )

            verdict = verdict_for_score(
                score,
                bulk_lot=bulk_lot,
                reissue=reissue,
            )

            connection.execute(
                upsert_statement,
                {
                    "marketplace": row["marketplace"],
                    "listing_id": row["listing_id"],
                    "source_fingerprint": row_fingerprint,

                    "auto_catalog_number": catalog_number,
                    "auto_region": region,
                    "auto_media_type": media_type,
                    "auto_disc_count": disc_count,
                    "auto_bulk_lot": bulk_lot,

                    "auto_obi": obi,
                    "auto_insert_present": insert_present,
                    "auto_poster_present": poster_present,
                    "auto_rental": rental,
                    "auto_sticker": sticker,
                    "auto_promo": promo,
                    "auto_sealed": sealed,
                    "auto_reissue": reissue,
                    "auto_first_press": first_press,

                    "seller_total_sales": seller.total_sales,
                    "seller_first_sale_at": seller.first_sale_at,
                    "seller_last_sale_at": seller.last_sale_at,
                    "seller_average_gross_price": (
                        seller.average_gross_price
                    ),
                    "repeat_seller": seller.repeat_seller,

                    "auction_duration_days": duration,
                    "start_to_finish_multiplier": (
                        price_multiplier
                    ),
                    "bids_per_day": bids_per_day,

                    "auto_importance_score": score,
                    "auto_verdict": verdict,
                },
            )

            changed += 1

    return scanned, changed


def export_query(
    marketplace: str,
) -> list[dict[str, Any]]:
    where_clause = ""

    if marketplace != "all":
        where_clause = "WHERE marketplace = :marketplace"

    statement = text(
        f"""
        SELECT
            marketplace,
            listing_id,
            ended_at,
            seller,
            title,

            effective_catalog_number,
            effective_region,
            effective_media_type,
            effective_disc_count,
            effective_bulk_lot,

            effective_obi,
            effective_insert_present,
            effective_poster_present,
            effective_rental,
            effective_sticker,
            effective_promo,
            effective_sealed,
            effective_reissue,
            effective_first_press,

            seller_total_sales,
            seller_first_sale_at,
            seller_last_sale_at,
            seller_average_gross_price,
            repeat_seller,

            auction_duration_days,
            start_to_finish_multiplier,
            bids_per_day,

            effective_importance_score,
            effective_verdict,

            effective_condition_media,
            effective_condition_cover,

            manual_catalog_number,
            manual_region,
            manual_media_type,
            manual_disc_count,
            manual_bulk_lot,
            manual_obi,
            manual_insert_present,
            manual_poster_present,
            manual_rental,
            manual_sticker,
            manual_promo,
            manual_sealed,
            manual_reissue,
            manual_first_press,
            manual_importance_score,
            manual_verdict,
            manual_condition_media,
            manual_condition_cover,
            manual_completeness_notes,
            manual_collector_notes,

            start_price,
            final_price,
            tax_amount,
            gross_price,
            currency,
            bid_count,
            auction_url

        FROM warehouse.auction_collector_effective
        {where_clause}
        ORDER BY
            ended_at DESC NULLS LAST,
            id DESC
        """
    )

    parameters = (
        {}
        if marketplace == "all"
        else {"marketplace": marketplace}
    )

    with engine.connect() as connection:
        rows = connection.execute(
            statement,
            parameters,
        ).mappings().all()

    return [dict(row) for row in rows]


def export_csv(
    rows: list[dict[str, Any]],
    path: Path,
) -> Path:
    if not rows:
        raise RuntimeError(
            "No collector rows were available."
        )

    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with path.open(
        "w",
        encoding="utf-8-sig",
        newline="",
    ) as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=list(rows[0]),
        )
        writer.writeheader()
        writer.writerows(rows)

    return path


def excel_cell_value(value: Any) -> Any:
    """Return a value that openpyxl can safely serialize."""
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            return value.isoformat()

        return value

    if isinstance(value, Decimal):
        return float(value)

    return value


def export_xlsx(
    rows: list[dict[str, Any]],
    path: Path,
) -> Path:
    if not rows:
        raise RuntimeError(
            "No collector rows were available."
        )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Collector Review"

    headers = list(rows[0])
    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        worksheet.append(
            [
                excel_cell_value(
                    row.get(header)
                )
                for header in headers
            ]
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for index, header in enumerate(
        headers,
        start=1,
    ):
        sample_values = [
            str(header),
            *[
                str(row.get(header) or "")
                for row in rows[:200]
            ],
        ]

        worksheet.column_dimensions[
            get_column_letter(index)
        ].width = min(
            max(
                len(value)
                for value in sample_values
            )
            + 2,
            55,
        )

    manual_headers = {
        index
        for index, header in enumerate(
            headers,
            start=1,
        )
        if header.startswith("manual_")
    }

    for column_index in manual_headers:
        worksheet.cell(
            row=1,
            column=column_index,
        ).font = Font(
            bold=True,
            italic=True,
        )

    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )
    workbook.save(path)

    return path


def export_review(
    marketplace: str,
    output_dir: Path,
) -> tuple[Path, Path]:
    create_schema()
    rows = export_query(marketplace)

    suffix = (
        ""
        if marketplace == "all"
        else f"_{marketplace}"
    )

    csv_path = export_csv(
        rows,
        output_dir
        / f"collector_review{suffix}.csv",
    )
    xlsx_path = export_xlsx(
        rows,
        output_dir
        / f"collector_review{suffix}.xlsx",
    )

    return csv_path, xlsx_path


def parse_optional_bool(
    value: Any,
) -> bool | None:
    if value is None:
        return None

    normalized = str(value).strip().casefold()

    if normalized == "":
        return None

    if normalized == CLEAR_VALUE.casefold():
        return None

    if normalized in {
        "1",
        "true",
        "yes",
        "y",
        "present",
    }:
        return True

    if normalized in {
        "0",
        "false",
        "no",
        "n",
        "missing",
        "absent",
    }:
        return False

    raise ValueError(
        f"Invalid boolean value: {value!r}"
    )


def parse_optional_int(
    value: Any,
) -> int | None:
    if value is None:
        return None

    normalized = str(value).strip()

    if normalized in {
        "",
        CLEAR_VALUE,
    }:
        return None

    return int(
        Decimal(normalized)
    )


def import_rows_from_csv(
    path: Path,
) -> list[dict[str, Any]]:
    with path.open(
        encoding="utf-8-sig",
        newline="",
    ) as handle:
        return list(
            csv.DictReader(handle)
        )


def import_rows_from_xlsx(
    path: Path,
) -> list[dict[str, Any]]:
    workbook = load_workbook(
        path,
        data_only=True,
    )
    worksheet = workbook.active

    values = list(
        worksheet.iter_rows(
            values_only=True,
        )
    )

    if not values:
        return []

    headers = [
        str(value)
        if value is not None
        else ""
        for value in values[0]
    ]

    return [
        dict(
            zip(
                headers,
                row,
                strict=False,
            )
        )
        for row in values[1:]
    ]


def import_review(
    path: Path,
) -> tuple[int, int]:
    create_schema()

    if not path.exists():
        raise FileNotFoundError(path)

    if path.suffix.casefold() == ".csv":
        rows = import_rows_from_csv(path)
    elif path.suffix.casefold() == ".xlsx":
        rows = import_rows_from_xlsx(path)
    else:
        raise ValueError(
            "Review import supports CSV or XLSX."
        )

    editable_columns = {
        "manual_catalog_number": "text",
        "manual_region": "text",
        "manual_media_type": "text",
        "manual_disc_count": "integer",
        "manual_bulk_lot": "boolean",
        "manual_obi": "boolean",
        "manual_insert_present": "boolean",
        "manual_poster_present": "boolean",
        "manual_rental": "boolean",
        "manual_sticker": "boolean",
        "manual_promo": "boolean",
        "manual_sealed": "boolean",
        "manual_reissue": "boolean",
        "manual_first_press": "boolean",
        "manual_importance_score": "integer",
        "manual_verdict": "text",
        "manual_condition_media": "text",
        "manual_condition_cover": "text",
        "manual_completeness_notes": "text",
        "manual_collector_notes": "text",
    }

    updated_rows = 0
    updated_fields = 0

    with engine.begin() as connection:
        for row in rows:
            marketplace = str(
                row.get("marketplace") or ""
            ).strip()
            listing_id = str(
                row.get("listing_id") or ""
            ).strip()

            if not marketplace or not listing_id:
                continue

            assignments: list[str] = []
            parameters: dict[str, Any] = {
                "marketplace": marketplace,
                "listing_id": listing_id,
            }

            for column, column_type in editable_columns.items():
                if column not in row:
                    continue

                raw_value = row.get(column)

                if raw_value is None:
                    continue

                normalized = str(raw_value).strip()

                if normalized == "":
                    continue

                if normalized == CLEAR_VALUE:
                    value = None
                elif column_type == "boolean":
                    value = parse_optional_bool(
                        raw_value
                    )
                elif column_type == "integer":
                    value = parse_optional_int(
                        raw_value
                    )
                else:
                    value = normalized

                assignments.append(
                    f"{column} = :{column}"
                )
                parameters[column] = value
                updated_fields += 1

            if not assignments:
                continue

            statement = text(
                f"""
                UPDATE warehouse.auction_collector
                SET
                    {", ".join(assignments)},
                    updated_at = NOW()
                WHERE marketplace = :marketplace
                  AND listing_id = :listing_id
                """
            )

            result = connection.execute(
                statement,
                parameters,
            )

            if result.rowcount:
                updated_rows += 1

    return updated_rows, updated_fields


def print_status() -> None:
    create_schema()

    statement = text(
        """
        SELECT
            marketplace,
            COUNT(*) AS rows,

            COUNT(effective_catalog_number)
                AS catalog_numbers,
            COUNT(effective_region)
                AS regions,
            COUNT(effective_media_type)
                AS media_types,
            COUNT(effective_disc_count)
                AS disc_counts,

            COUNT(*) FILTER (
                WHERE effective_obi IS NOT NULL
            ) AS obi_values,

            COUNT(*) FILTER (
                WHERE effective_insert_present IS NOT NULL
            ) AS insert_values,

            COUNT(*) FILTER (
                WHERE effective_poster_present IS NOT NULL
            ) AS poster_values,

            COUNT(auction_duration_days)
                AS durations,
            COUNT(start_to_finish_multiplier)
                AS multipliers,
            COUNT(bids_per_day)
                AS bid_velocities,

            COUNT(effective_importance_score)
                AS importance_scores,
            COUNT(effective_verdict)
                AS verdicts,

            COUNT(*) FILTER (
                WHERE manual_collector_notes IS NOT NULL
            ) AS manually_reviewed

        FROM warehouse.auction_collector_effective
        GROUP BY marketplace
        ORDER BY marketplace
        """
    )

    duplicate_statement = text(
        """
        SELECT COUNT(*)
        FROM (
            SELECT
                marketplace,
                listing_id
            FROM warehouse.auction_collector
            GROUP BY
                marketplace,
                listing_id
            HAVING COUNT(*) > 1
        ) AS duplicates
        """
    )

    with engine.connect() as connection:
        rows = connection.execute(
            statement
        ).mappings().all()

        duplicates = connection.execute(
            duplicate_statement
        ).scalar_one()

    print()
    print("Collector feature coverage")
    print("--------------------------")

    for row in rows:
        print()
        print(row["marketplace"])
        print(
            f"  Rows              : {row['rows']}"
        )
        print(
            f"  Catalog numbers   : {row['catalog_numbers']}"
        )
        print(
            f"  Regions           : {row['regions']}"
        )
        print(
            f"  Media types       : {row['media_types']}"
        )
        print(
            f"  Disc counts       : {row['disc_counts']}"
        )
        print(
            f"  Obi values        : {row['obi_values']}"
        )
        print(
            f"  Insert values     : {row['insert_values']}"
        )
        print(
            f"  Poster values     : {row['poster_values']}"
        )
        print(
            f"  Auction durations : {row['durations']}"
        )
        print(
            f"  Price multipliers : {row['multipliers']}"
        )
        print(
            f"  Bid velocities    : {row['bid_velocities']}"
        )
        print(
            f"  Importance scores : {row['importance_scores']}"
        )
        print(
            f"  Verdicts          : {row['verdicts']}"
        )
        print(
            f"  Manual reviews    : {row['manually_reviewed']}"
        )

    print()
    print(
        "Duplicate collector rows:",
        duplicates,
    )


def main() -> int:
    args = parse_args()

    if args.command == "build":
        scanned, changed = build_features(
            marketplace=args.marketplace,
            force=args.force,
        )

        print()
        print("Collector feature build")
        print("-----------------------")
        print("Scanned :", scanned)
        print("Changed :", changed)
        return 0

    if args.command == "export":
        csv_path, xlsx_path = export_review(
            marketplace=args.marketplace,
            output_dir=args.output_dir,
        )

        print()
        print("Collector review export")
        print("-----------------------")
        print("CSV :", csv_path)
        print("XLSX:", xlsx_path)
        print(
            "Use __CLEAR__ to explicitly clear a manual value."
        )
        return 0

    if args.command == "import":
        updated_rows, updated_fields = import_review(
            args.path
        )

        print()
        print("Collector review import")
        print("-----------------------")
        print("Rows updated  :", updated_rows)
        print("Fields updated:", updated_fields)
        return 0

    if args.command == "status":
        print_status()
        return 0

    raise RuntimeError(
        f"Unsupported command: {args.command}"
    )


if __name__ == "__main__":
    raise SystemExit(main())
