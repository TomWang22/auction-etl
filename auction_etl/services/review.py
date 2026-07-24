from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from auction_etl.models.staging import Listing


EDITABLE_FIELDS = (
    "format",
    "disc_count",
    "catalog_number",
    "label",
    "year",
    "country",
    "media_condition",
    "sleeve_condition",
    "obi",
    "edition",
    "description",
    "condition_text",
)

REVIEW_COLUMNS = (
    "marketplace",
    "listing_id",
    "title",
    "auction_url",
    "seller",
    *EDITABLE_FIELDS,
    "review_notes",
)

CLEAR_VALUE = "__CLEAR__"


@dataclass(slots=True)
class ReviewExportStats:
    rows: int
    path: Path


@dataclass(slots=True)
class ReviewImportStats:
    scanned: int = 0
    matched: int = 0
    updated_rows: int = 0
    updated_fields: int = 0
    missing: int = 0
    invalid: int = 0


def _serialize(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, bool):
        return "true" if value else "false"

    return str(value)


def _needs_review_expression():
    return or_(
        Listing.format.is_(None),
        Listing.catalog_number.is_(None),
        Listing.label.is_(None),
        Listing.media_condition.is_(None),
        Listing.sleeve_condition.is_(None),
        Listing.obi.is_(None),
        Listing.edition.is_(None),
    )


def export_review_csv(
    session: Session,
    output_path: str | Path,
    *,
    marketplace: str | None = None,
    all_rows: bool = False,
) -> ReviewExportStats:
    statement = select(Listing).order_by(
        Listing.marketplace,
        Listing.id,
    )

    if marketplace:
        statement = statement.where(
            Listing.marketplace == marketplace
        )

    if not all_rows:
        statement = statement.where(
            _needs_review_expression()
        )

    listings = session.scalars(statement).all()
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    with path.open(
        "w",
        encoding="utf-8-sig",
        newline="",
    ) as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=list(REVIEW_COLUMNS),
        )
        writer.writeheader()

        for listing in listings:
            row = {
                "marketplace": listing.marketplace,
                "listing_id": listing.listing_id,
                "title": listing.title or "",
                "auction_url": listing.auction_url,
                "seller": listing.seller or "",
                "review_notes": "",
            }

            for field in EDITABLE_FIELDS:
                row[field] = _serialize(
                    getattr(listing, field)
                )

            writer.writerow(row)

    return ReviewExportStats(
        rows=len(listings),
        path=path,
    )


def _parse_bool(value: str) -> bool:
    normalized = value.strip().casefold()

    if normalized in {
        "true",
        "yes",
        "y",
        "1",
        "present",
        "included",
    }:
        return True

    if normalized in {
        "false",
        "no",
        "n",
        "0",
        "none",
        "absent",
        "missing",
        "not included",
    }:
        return False

    raise ValueError(f"Invalid boolean: {value}")


def _parse_value(
    field: str,
    raw_value: str,
) -> Any:
    value = raw_value.strip()

    if value == CLEAR_VALUE:
        return None

    if field == "disc_count":
        disc_count = int(value)

        if disc_count < 1 or disc_count > 500:
            raise ValueError(
                f"Invalid disc count: {disc_count}"
            )

        return disc_count

    if field == "year":
        year = int(value)

        if year < 1900 or year > 2100:
            raise ValueError(
                f"Invalid year: {year}"
            )

        return year

    if field == "obi":
        return _parse_bool(value)

    return value


def import_review_csv(
    session: Session,
    input_path: str | Path,
) -> ReviewImportStats:
    path = Path(input_path)

    if not path.exists():
        raise FileNotFoundError(path)

    stats = ReviewImportStats()

    with path.open(
        "r",
        encoding="utf-8-sig",
        newline="",
    ) as handle:
        reader = csv.DictReader(handle)

        required = {
            "marketplace",
            "listing_id",
        }

        missing_columns = required.difference(
            reader.fieldnames or []
        )

        if missing_columns:
            raise ValueError(
                "Missing required columns: "
                + ", ".join(
                    sorted(missing_columns)
                )
            )

        for row in reader:
            stats.scanned += 1

            marketplace = (
                row.get("marketplace") or ""
            ).strip()
            listing_id = (
                row.get("listing_id") or ""
            ).strip()

            if not marketplace or not listing_id:
                stats.invalid += 1
                continue

            listing = session.scalar(
                select(Listing).where(
                    Listing.marketplace == marketplace,
                    Listing.listing_id == listing_id,
                )
            )

            if listing is None:
                stats.missing += 1
                continue

            stats.matched += 1
            row_changed = False

            for field in EDITABLE_FIELDS:
                raw_value = row.get(field)

                if raw_value is None:
                    continue

                raw_value = raw_value.strip()

                if raw_value == "":
                    continue

                try:
                    value = _parse_value(
                        field,
                        raw_value,
                    )
                except ValueError:
                    stats.invalid += 1
                    continue

                if getattr(listing, field) == value:
                    continue

                setattr(listing, field, value)
                row_changed = True
                stats.updated_fields += 1

            notes = (
                row.get("review_notes") or ""
            ).strip()

            if notes:
                payload = dict(listing.payload or {})
                review = dict(
                    payload.get("manual_review") or {}
                )
                review["notes"] = notes
                payload["manual_review"] = review
                listing.payload = payload
                row_changed = True
                stats.updated_fields += 1

            if row_changed:
                stats.updated_rows += 1

    session.commit()
    return stats


def review_coverage(
    session: Session,
) -> dict[str, tuple[int, int, float]]:
    total = session.scalar(
        select(func.count()).select_from(Listing)
    ) or 0

    coverage: dict[
        str,
        tuple[int, int, float],
    ] = {}

    for field in EDITABLE_FIELDS:
        column = getattr(Listing, field)

        populated = session.scalar(
            select(func.count())
            .select_from(Listing)
            .where(column.is_not(None))
        ) or 0

        percent = (
            populated / total * 100
            if total
            else 0.0
        )

        coverage[field] = (
            populated,
            total,
            percent,
        )

    return coverage


def export_review_xlsx(
    session: Session,
    output_path: str | Path,
    *,
    marketplace: str | None = None,
    all_rows: bool = False,
) -> ReviewExportStats:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table
    from openpyxl.worksheet.table import TableStyleInfo

    statement = select(Listing).order_by(
        Listing.marketplace,
        Listing.id,
    )

    if marketplace:
        statement = statement.where(
            Listing.marketplace == marketplace
        )

    if not all_rows:
        statement = statement.where(
            _needs_review_expression()
        )

    listings = session.scalars(
        statement
    ).all()

    path = Path(output_path)
    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Review"

    sheet.append(
        list(REVIEW_COLUMNS)
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    for cell in sheet[1]:
        cell.font = Font(
            bold=True
        )
        cell.fill = header_fill

    for listing in listings:
        row = {
            "marketplace": (
                listing.marketplace
            ),
            "listing_id": (
                listing.listing_id
            ),
            "title": (
                listing.title or ""
            ),
            "auction_url": (
                listing.auction_url
            ),
            "seller": (
                listing.seller or ""
            ),
            "review_notes": "",
        }

        for field in EDITABLE_FIELDS:
            row[field] = _serialize(
                getattr(
                    listing,
                    field,
                )
            )

        sheet.append(
            [
                row[column]
                for column
                in REVIEW_COLUMNS
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = (
        sheet.dimensions
    )

    if listings:
        table = Table(
            displayName="ListingReview",
            ref=sheet.dimensions,
        )
        table.tableStyleInfo = (
            TableStyleInfo(
                name=(
                    "TableStyleMedium2"
                ),
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
        )
        sheet.add_table(table)

    widths = {
        "marketplace": 12,
        "listing_id": 18,
        "title": 70,
        "auction_url": 65,
        "seller": 28,
        "format": 20,
        "disc_count": 12,
        "catalog_number": 20,
        "label": 24,
        "year": 10,
        "country": 20,
        "media_condition": 18,
        "sleeve_condition": 18,
        "obi": 12,
        "edition": 24,
        "description": 60,
        "condition_text": 30,
        "review_notes": 50,
    }

    for index, column in enumerate(
        REVIEW_COLUMNS,
        start=1,
    ):
        sheet.column_dimensions[
            get_column_letter(index)
        ].width = widths.get(
            column,
            18,
        )

    workbook.save(path)

    return ReviewExportStats(
        rows=len(listings),
        path=path,
    )


def import_review_xlsx(
    session: Session,
    input_path: str | Path,
) -> ReviewImportStats:
    from openpyxl import load_workbook

    path = Path(input_path)

    if not path.exists():
        raise FileNotFoundError(path)

    workbook = load_workbook(
        path,
        data_only=True,
    )
    sheet = workbook["Review"]

    rows = list(
        sheet.iter_rows(
            values_only=True,
        )
    )

    if not rows:
        return ReviewImportStats()

    headers = [
        str(value or "").strip()
        for value in rows[0]
    ]

    required = {
        "marketplace",
        "listing_id",
    }

    missing_columns = (
        required.difference(headers)
    )

    if missing_columns:
        raise ValueError(
            "Missing required columns: "
            + ", ".join(
                sorted(
                    missing_columns
                )
            )
        )

    stats = ReviewImportStats()

    for values in rows[1:]:
        row = {
            header: (
                "" if value is None
                else str(value)
            )
            for header, value
            in zip(
                headers,
                values,
                strict=False,
            )
        }

        stats.scanned += 1

        marketplace = (
            row.get(
                "marketplace",
                "",
            ).strip()
        )
        listing_id = (
            row.get(
                "listing_id",
                "",
            ).strip()
        )

        if (
            not marketplace
            or not listing_id
        ):
            stats.invalid += 1
            continue

        listing = session.scalar(
            select(Listing).where(
                Listing.marketplace
                == marketplace,
                Listing.listing_id
                == listing_id,
            )
        )

        if listing is None:
            stats.missing += 1
            continue

        stats.matched += 1
        row_changed = False

        for field in EDITABLE_FIELDS:
            raw_value = row.get(
                field,
                "",
            ).strip()

            if raw_value == "":
                continue

            try:
                value = _parse_value(
                    field,
                    raw_value,
                )
            except ValueError:
                stats.invalid += 1
                continue

            if (
                getattr(
                    listing,
                    field,
                )
                == value
            ):
                continue

            setattr(
                listing,
                field,
                value,
            )
            row_changed = True
            stats.updated_fields += 1

        notes = row.get(
            "review_notes",
            "",
        ).strip()

        if notes:
            payload = dict(
                listing.payload or {}
            )
            review = dict(
                payload.get(
                    "manual_review"
                )
                or {}
            )
            review["notes"] = notes
            payload[
                "manual_review"
            ] = review
            listing.payload = payload
            row_changed = True
            stats.updated_fields += 1

        if row_changed:
            stats.updated_rows += 1

    session.commit()
    return stats
